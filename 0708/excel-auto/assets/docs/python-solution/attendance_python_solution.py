#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WEEKDAY_ZH = "一二三四五六日"
EMPLOYEE_RE = re.compile(r"^(?:員工代號[:：]\s*)?(\d{2,})\s+(.+)$")
DATE_TEXT_RE = re.compile(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})")
TIME_AT_END_RE = re.compile(r"(\d{1,2})[:：](\d{2})\s*$")


@dataclass(frozen=True)
class ColumnMap:
    date_col: int
    weekday_col: int | None
    shift_code_col: int | None
    shift_name_col: int | None
    start_col: int | None
    end_col: int | None
    note_col: int | None


@dataclass(frozen=True)
class AttendanceRecord:
    source_row: int
    employee_raw: str
    employee_masked: str
    work_date: date
    weekday: str
    shift_code: object
    shift_name: object
    start_raw: object
    end_raw: object
    note: object
    start_minutes: int | None
    end_minutes: int | None
    capped_end_minutes: int | None
    work_minutes: int
    day_count: int


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_employee(value: object) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    match = EMPLOYEE_RE.fullmatch(text)
    if not match:
        return None
    return f"{match.group(1)} {match.group(2).strip()}"


def mask_employee(raw: str) -> str:
    match = EMPLOYEE_RE.fullmatch(normalize_text(raw))
    if not match:
        return normalize_text(raw)
    emp_id, name = match.group(1), match.group(2).strip()
    masked_id = (
        "*" * len(emp_id)
        if len(emp_id) <= 2
        else f"{emp_id[:2]}{'*' * max(1, len(emp_id) - 4)}{emp_id[-2:]}"
    )
    chars = list(name)
    if len(chars) <= 1:
        masked_name = "*"
    elif len(chars) == 2:
        masked_name = f"{chars[0]}*"
    else:
        masked_name = f"{chars[0]}{'*' * (len(chars) - 2)}{chars[-1]}"
    return f"員工代號:{masked_id} {masked_name}"


def parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    match = DATE_TEXT_RE.fullmatch(text)
    if not match:
        return None
    year, month, day = (int(part) for part in match.groups())
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_time_minutes(value: object) -> int | None:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, (int, float)) and 0 <= value < 1:
        return round(float(value) * 24 * 60)
    text = normalize_text(value)
    match = TIME_AT_END_RE.search(text)
    if not match:
        return None
    hour, minute = int(match.group(1)), int(match.group(2))
    if hour > 23 or minute > 59:
        return None
    return hour * 60 + minute


def format_hhmm(minutes: int | None) -> str:
    if minutes is None:
        return ""
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def format_duration(minutes: int) -> str:
    return f"{minutes // 60}:{minutes % 60:02d}"


def detect_columns(header_values: Iterable[object]) -> ColumnMap | None:
    labels = {normalize_text(value): idx for idx, value in enumerate(header_values, start=1)}
    aliases = {
        "date_col": ("日期", "出勤日期", "工作日期"),
        "weekday_col": ("星期", "週", "周"),
        "shift_code_col": ("班次代號", "班別代號"),
        "shift_name_col": ("班次名稱", "班別名稱"),
        "start_col": ("上班", "上班時間", "到班", "刷入"),
        "end_col": ("下班", "下班時間", "退班", "刷出"),
        "note_col": ("備註", "說明"),
    }

    found: dict[str, int | None] = {}
    for key, names in aliases.items():
        found[key] = next((labels[name] for name in names if name in labels), None)

    if found["date_col"] is None:
        return None
    return ColumnMap(**found)


def fallback_columns() -> ColumnMap:
    return ColumnMap(
        date_col=1,
        weekday_col=2,
        shift_code_col=3,
        shift_name_col=4,
        start_col=5,
        end_col=6,
        note_col=7,
    )


def row_value(row: tuple[object, ...], col: int | None) -> object:
    if col is None or col < 1 or col > len(row):
        return None
    return row[col - 1]


def score_sheet(ws) -> int:
    score = 0
    for row in ws.iter_rows(values_only=True):
        first = row[0] if row else None
        if parse_employee(first):
            score += 5
        if parse_date(first):
            score += 1
        if detect_columns(row):
            score += 3
    return score


def choose_attendance_sheet(workbook, requested_name: str | None):
    if requested_name:
        if requested_name not in workbook.sheetnames:
            raise ValueError(f"找不到工作表：{requested_name}")
        return workbook[requested_name]

    preferred_names = ("員工出勤明細表", "出勤明細", "出勤資料", "attendance")
    for name in preferred_names:
        for sheet_name in workbook.sheetnames:
            if name.lower() in sheet_name.lower():
                return workbook[sheet_name]

    scored = sorted(((score_sheet(ws), ws) for ws in workbook.worksheets), key=lambda item: item[0], reverse=True)
    if not scored or scored[0][0] == 0:
        raise ValueError("無法辨識出勤明細工作表，請使用 --sheet 指定。")
    return scored[0][1]


def parse_cap(value: str) -> int:
    parsed = parse_time_minutes(value)
    if parsed is None:
        raise argparse.ArgumentTypeError("時間上限必須是 HH:MM，例如 19:30")
    return parsed


def weekday_for(work_date: date, explicit: object) -> str:
    text = normalize_text(explicit)
    if text in set(WEEKDAY_ZH):
        return text
    return WEEKDAY_ZH[work_date.weekday()]


def parse_records(ws, cap_minutes: int) -> list[AttendanceRecord]:
    current_employee = ""
    current_columns = fallback_columns()
    records: list[AttendanceRecord] = []
    counted_days: set[tuple[str, date]] = set()

    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        detected_columns = detect_columns(row)
        if detected_columns:
            current_columns = detected_columns
            continue

        first_cell = row[0] if row else None
        employee = parse_employee(first_cell)
        if employee:
            current_employee = employee
            continue

        work_date = parse_date(row_value(row, current_columns.date_col))
        if work_date is None or not current_employee:
            continue

        start_raw = row_value(row, current_columns.start_col)
        end_raw = row_value(row, current_columns.end_col)
        start_minutes = parse_time_minutes(start_raw)
        end_minutes = parse_time_minutes(end_raw)
        capped_end_minutes = min(end_minutes, cap_minutes) if end_minutes is not None else None
        weekday = weekday_for(work_date, row_value(row, current_columns.weekday_col))
        is_weekend = weekday in {"六", "日"} or work_date.weekday() >= 5

        if start_minutes is None or capped_end_minutes is None or is_weekend:
            work_minutes = 0
        else:
            work_minutes = max(0, capped_end_minutes - start_minutes)

        day_key = (current_employee, work_date)
        day_count = 1 if work_minutes > 0 and day_key not in counted_days else 0
        if day_count:
            counted_days.add(day_key)

        records.append(
            AttendanceRecord(
                source_row=row_index,
                employee_raw=current_employee,
                employee_masked=mask_employee(current_employee),
                work_date=work_date,
                weekday=weekday,
                shift_code=row_value(row, current_columns.shift_code_col),
                shift_name=row_value(row, current_columns.shift_name_col),
                start_raw=start_raw,
                end_raw=end_raw,
                note=row_value(row, current_columns.note_col),
                start_minutes=start_minutes,
                end_minutes=end_minutes,
                capped_end_minutes=capped_end_minutes,
                work_minutes=work_minutes,
                day_count=day_count,
            )
        )

    if not records:
        raise ValueError("沒有解析到任何出勤日期列，請確認工作表結構。")
    return records


def summarize(records: list[AttendanceRecord]) -> OrderedDict[tuple[str, str], dict[str, int]]:
    summary: OrderedDict[tuple[str, str], dict[str, int]] = OrderedDict()
    for record in records:
        month_key = record.work_date.strftime("%Y-%m")
        key = (record.employee_masked, month_key)
        if key not in summary:
            summary[key] = {"workdays": 0, "minutes": 0}
        summary[key]["workdays"] += record.day_count
        summary[key]["minutes"] += record.work_minutes
    return summary


def write_output(
    input_path: Path,
    output_path: Path,
    source_sheet: str,
    records: list[AttendanceRecord],
    cap_minutes: int,
) -> None:
    workbook = load_workbook(input_path)
    if "算工作數" not in workbook.sheetnames:
        raise ValueError("找不到工作表：算工作數")
    mask_employee_headers(workbook)
    summary_ws = workbook["算工作數"]

    header_row = find_summary_header_row(summary_ws)
    clear_summary_rows(summary_ws, header_row + 1)
    for (employee, month_key), values in summarize(records).items():
        summary_ws.append([employee, month_key, values["workdays"], format_duration(values["minutes"])])

    if "計算明細" in workbook.sheetnames:
        del workbook["計算明細"]
    detail_ws = workbook.create_sheet("計算明細")
    write_audit_detail_sheet(detail_ws, records)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def find_summary_header_row(ws) -> int:
    for row in range(1, min(ws.max_row, 20) + 1):
        values = [ws.cell(row, col).value for col in range(1, 4)]
        if values == ["每月", "工作天數", "工作總時數"]:
            ws.cell(row, 1).value = "員工"
            ws.cell(row, 2).value = "每月"
            ws.cell(row, 3).value = "工作天數"
            ws.cell(row, 4).value = "工作總時數"
            return row
        if values == ["員工", "每月", "工作天數"]:
            ws.cell(row, 4).value = "工作總時數"
            return row
    raise ValueError("無法在算工作數工作表中找到彙總表頭。")


def clear_summary_rows(ws, start_row: int) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_audit_detail_sheet(ws, records: list[AttendanceRecord]) -> None:
    ws.append(
        [
            "員工",
            "日期",
            "星期",
            "班次代號",
            "班次名稱",
            "上班原始",
            "下班原始",
            "上班時間",
            "下班時間",
            "計算下班",
            "工作時數",
            "每日計數",
            "來源列",
            "備註",
        ]
    )
    for record in records:
        ws.append(
            [
                record.employee_masked,
                record.work_date,
                record.weekday,
                record.shift_code,
                record.shift_name,
                record.start_raw,
                record.end_raw,
                format_hhmm(record.start_minutes),
                format_hhmm(record.end_minutes),
                format_hhmm(record.capped_end_minutes),
                format_duration(record.work_minutes) if record.work_minutes else "",
                record.day_count,
                record.source_row,
                record.note,
            ]
        )
    style_header(ws)
    autosize_columns(ws)
    ws.freeze_panes = "A2"
    for cell in ws["B"][1:]:
        cell.number_format = "yyyy-mm-dd"


def mask_employee_headers(workbook) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            employee = parse_employee(cell.value)
            if employee:
                cell.value = mask_employee(employee)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        width = 10
        for cell in column_cells:
            value = cell.value
            if isinstance(value, date):
                length = 10
            else:
                length = len(str(value)) if value is not None else 0
            width = max(width, min(length + 2, 42))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}-python工時計算.xlsx")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="計算員工出勤工作天數與工作總時數。")
    parser.add_argument("input", type=Path, help="來源 Excel 檔案，例如 data/2026.06月份.xlsx")
    parser.add_argument("-o", "--output", type=Path, help="輸出 Excel 檔案路徑")
    parser.add_argument("--sheet", help="指定出勤明細工作表名稱；未指定時會自動偵測")
    parser.add_argument("--cap", type=parse_cap, default=parse_cap("19:30"), help="下班時間上限，預設 19:30")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    input_path = args.input
    output_path = args.output or default_output_path(input_path)

    workbook = load_workbook(input_path, data_only=True)
    source_ws = choose_attendance_sheet(workbook, args.sheet)
    records = parse_records(source_ws, args.cap)
    write_output(input_path, output_path, source_ws.title, records, args.cap)

    print(f"source_sheet={source_ws.title}")
    print(f"records={len(records)}")
    print(f"employees={len({record.employee_masked for record in records})}")
    print(f"output={output_path}")


if __name__ == "__main__":
    main()
